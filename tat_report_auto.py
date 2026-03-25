"""
Valuecart Seller Queries TAT Report — Automated Weekly Script
Runs every Thursday via GitHub Actions.
Fetches tickets from Zoho Desk API, builds Excel report, sends via Gmail.
"""

import os, io, base64, requests, pandas as pd, smtplib
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.formula import ArrayFormula
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ── Secrets ───────────────────────────────────────────────────
ZOHO_CLIENT_ID     = os.environ["ZOHO_CLIENT_ID"]
ZOHO_CLIENT_SECRET = os.environ["ZOHO_CLIENT_SECRET"]
ZOHO_REFRESH_TOKEN = os.environ["ZOHO_REFRESH_TOKEN"]
ZOHO_ORG_ID        = os.environ["ZOHO_ORG_ID"]
GMAIL_ADDRESS      = os.environ["GMAIL_ADDRESS"]
GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]
EMAIL_TO           = os.environ["EMAIL_TO"]
EMAIL_CC           = os.environ.get("EMAIL_CC", "")

# ── Week ranges (Sun–Sat) ─────────────────────────────────────
def week_num(d):
    jan1 = datetime(d.year, 1, 1)
    dow  = (jan1.weekday() + 1) % 7
    return ((d - jan1).days + dow) // 7 + 1

def week_bounds(d):
    dow = (d.weekday() + 1) % 7
    sun = d - timedelta(days=dow)
    return sun.replace(hour=0,minute=0,second=0,microsecond=0),            (sun + timedelta(days=6)).replace(hour=23,minute=59,second=59)

today          = datetime.now()
curr_sun, _    = week_bounds(today)
wB_start, wB_end = week_bounds(curr_sun - timedelta(weeks=1))
wA_start, wA_end = week_bounds(curr_sun - timedelta(weeks=2))
WK_A = f"Week {week_num(wA_start)}"
WK_B = f"Week {week_num(wB_start)}"

print(f"Report: {WK_B} ({wB_start.date()} to {wB_end.date()}) vs {WK_A} ({wA_start.date()} to {wA_end.date()})")

# ── Zoho API ──────────────────────────────────────────────────
def zoho_token():
    r = requests.post("https://accounts.zoho.in/oauth/v2/token", params={
        "refresh_token": ZOHO_REFRESH_TOKEN,
        "client_id":     ZOHO_CLIENT_ID,
        "client_secret": ZOHO_CLIENT_SECRET,
        "grant_type":    "refresh_token",
    })
    d = r.json()
    if "access_token" not in d:
        raise Exception(f"Zoho auth error: {d}")
    return d["access_token"]

def parse_dt(s):
    if not s: return None
    try: return datetime.strptime(s.replace("T"," ").replace("Z","").replace(".000","")[:19], "%Y-%m-%d %H:%M:%S")
    except: return None

def fetch_ticket_detail(headers, ticket_id):
    """Get full ticket details including custom fields."""
    r = requests.get(f"https://desk.zoho.in/api/v1/tickets/{ticket_id}", headers=headers)
    if r.status_code == 200:
        return r.json()
    return {}

def fetch_tickets(token, from_dt, to_dt):
    """Fetch tickets using Zoho search API with date range filter."""
    headers = {"Authorization": f"Zoho-oauthtoken {token}", "orgId": ZOHO_ORG_ID}
    result, offset = [], 0

    from_str = from_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    to_str   = to_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")

    while True:
        # Use Zoho search API with createdTime filter
        r = requests.get("https://desk.zoho.in/api/v1/tickets/search",
                         headers=headers,
                         params={
                             "limit":          50,
                             "from":           offset,
                             "createdTime":    from_str,
                             "endCreatedTime": to_str,
                         })
        print(f"  Search API status: {r.status_code} (offset={offset})")
        if r.status_code != 200:
            print(f"  Search error: {r.text[:300]}")
            break
        data = r.json().get("data", [])
        print(f"  Got {len(data)} tickets")
        if not data:
            break

        # Fetch full details for each ticket to get custom fields
        for t in data:
            detail = fetch_ticket_detail(headers, t.get("id",""))
            if detail:
                result.append(detail)
            else:
                result.append(t)

        if len(data) < 50:
            break
        offset += 50

    print(f"  Total: {len(result)} tickets for {from_dt.date()} – {to_dt.date()}")
    if result:
        cf = result[0].get("cf") or {}
        print(f"  Sample cf keys: {list(cf.keys())[:5]}")
        print(f"  slaViolationType: {result[0].get('slaViolationType','')}")
    return result

    # Filter by date
    result = []
    for t in all_tix:
        ct = t.get("createdTime", "") or ""
        ct_clean = ct.replace("T"," ").replace("Z","").replace(".000","")[:19]
        try:
            ct_dt = datetime.strptime(ct_clean, "%Y-%m-%d %H:%M:%S")
            if from_dt <= ct_dt <= to_dt:
                result.append(t)
        except:
            pass
    print(f"  Filtered to {len(result)} tickets for {from_dt.date()} – {to_dt.date()}")
    return result

def get_ticket_detail(token, ticket_id):
    """Fetch single ticket with all custom fields."""
    headers = {"Authorization": f"Zoho-oauthtoken {token}", "orgId": ZOHO_ORG_ID}
    r = requests.get(f"https://desk.zoho.in/api/v1/tickets/{ticket_id}", headers=headers)
    if r.status_code == 200:
        return r.json()
    return {}

def to_df(tickets, token=None):
    if not tickets:
        return pd.DataFrame(columns=["SLA Violation Type","Subject","Status",
            "Ticket Closed Time","Request Id","Due Date","BD POC",
            "Query Category","Seller Category","Internal Department"])
    rows = []
    for t in tickets:
        cf  = t.get("cf") or {}
        # If cf is empty and we have token, fetch individual ticket detail
        if not cf and token:
            detail = get_ticket_detail(token, t.get("id",""))
            cf = detail.get("cf") or {}
            t  = {**t, **detail}
        sla = (t.get("slaViolationType") or cf.get("cf_sla_violation_type") or "Not Violated")
        rows.append({
            "SLA Violation Type":  sla,
            "Subject":             t.get("subject",""),
            "Status":              t.get("status",""),
            "Ticket Closed Time":  (t.get("closedTime") or ""),
            "Request Id":          (t.get("ticketNumber") or t.get("id","")),
            "Due Date":            (t.get("dueDate") or ""),
            "BD POC":              (t.get("assignee") or {}).get("name",""),
            "Query Category":      (cf.get("cf_query_category") or t.get("category","") or ""),
            "Seller Category":     (cf.get("cf_seller_category") or ""),
            "Internal Department": (cf.get("cf_internal_department") or ""),
        })
    return pd.DataFrame(rows)

print("Fetching tickets from Zoho...")
token   = zoho_token()
tix_A   = fetch_tickets(token, wA_start, wA_end)
tix_B   = fetch_tickets(token, wB_start, wB_end)
df_A    = to_df(tix_A, token)
df_B    = to_df(tix_B, token)
print(f"  {WK_A}: {len(df_A)} tickets")
print(f"  {WK_B}: {len(df_B)} tickets")

excl_df = pd.concat([df_A, df_B], ignore_index=True)
excl_df = excl_df[excl_df["Status"] != "On Hold"].reset_index(drop=True)
resv_df = excl_df[excl_df["SLA Violation Type"] == "Resolution Violation"].reset_index(drop=True)
print(f"  Excluding On Hold: {len(excl_df)}, Resolution Violation: {len(resv_df)}")

# ── Build Excel ───────────────────────────────────────────────
PURPLE = "FF7030A0"; LAVEND = "FFF2E6FF"

def sc(cell, val=None, bold=False, bg=None, fc=None, ha="general", sz=11, fmt=None):
    if val is not None: cell.value = val
    cell.font      = Font(bold=bold, color=fc or "FF000000", size=sz, name="Calibri")
    cell.alignment = Alignment(horizontal=ha, vertical="center")
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    if fmt: cell.number_format = fmt

def write_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 17.7; ws.column_dimensions["B"].width = 59.2
    ws.column_dimensions["C"].width = 7.6;  ws.column_dimensions["D"].width = 16.7
    ws.column_dimensions["E"].width = 9.9;  ws.column_dimensions["F"].width = 15.4
    ws.column_dimensions["G"].width = 11.8; ws.column_dimensions["H"].width = 21.4
    ws.column_dimensions["I"].width = 13.7; ws.column_dimensions["J"].width = 18.3
    for ci, col in enumerate(df.columns, 1):
        sc(ws.cell(1, ci), col, bold=True, ha="left" if ci==1 else "center")
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(df.columns, 1):
            v = row[col]; v = None if pd.isna(v) else v
            sc(ws.cell(ri, ci), v, ha="left" if ci==1 else "center")

wb = Workbook(); wb.remove(wb.active)
wA_num = week_num(wA_start); wB_num = week_num(wB_start)
write_sheet(wb, f"WK_{wA_num}", df_A)
write_sheet(wb, f"WK_{wB_num}", df_B)
write_sheet(wb, "Excluding On Hold", excl_df)
write_sheet(wb, "Resolution Violation", resv_df)

# Report sheet
ws  = wb.create_sheet("Report")
ws.column_dimensions["A"].width = 40.8; ws.column_dimensions["B"].width = 22.2
ws.column_dimensions["C"].width = 13.0; ws.column_dimensions["D"].width = 27.8
ws.column_dimensions["E"].width = 14.8; ws.row_dimensions[1].height = 28
WA  = f"WK_{wA_num}"; WB = f"WK_{wB_num}"; EX = "'Excluding On Hold'"

ws.merge_cells("A1:D1")
sc(ws["A1"], "Valuecart Seller Queries TAT Report", bold=True, bg=PURPLE, fc="FFFFFFFF", ha="center", sz=14)
for c in ["B1","C1","D1"]: ws[c].fill = PatternFill("solid", fgColor=PURPLE)

for coord, val, ha in [("A3","KPI","left"),("B3",WK_B,"center"),("C3",WK_A,"center"),("D3","Trend","center")]:
    sc(ws[coord], val, bold=True, bg=PURPLE, fc="FFFFFFFF", ha=ha)

kpi = [
    (4,"Total Queries Received",
     f'=COUNTIFS(\'{WB}\'!C2:C1000,"<>",\'{WB}\'!C2:C1000,"<>On Hold")',
     f'=COUNTIFS(\'{WA}\'!C2:C1000,"<>",\'{WA}\'!C2:C1000,"<>On Hold")',
     '=IF(B4>C4,"⬆",IF(B4<C4,"⬇","➡"))',None),
    (5,"Queries Closed",
     f'=COUNTIFS(\'{WB}\'!C2:C1000,"Closed")',
     f'=COUNTIFS(\'{WA}\'!C2:C1000,"Closed")',
     '=IF(B5>C5,"⬆",IF(B5<C5,"⬇","➡"))',None),
    (6,"Closure Within TAT (%)",
     f'=IFERROR(COUNTIFS(\'{WB}\'!A2:A1000,"Not Violated",\'{WB}\'!C2:C1000,"Closed")/COUNTIF(\'{WB}\'!C2:C1000,"Closed"),0)',
     f'=IFERROR(COUNTIFS(\'{WA}\'!A2:A1000,"Not Violated",\'{WA}\'!C2:C1000,"Closed")/COUNTIF(\'{WA}\'!C2:C1000,"Closed"),0)',
     '=IF(B6>C6,"⬆",IF(B6<C6,"⬇","➡"))',LAVEND),
    (7,"Open Queries",
     f'=COUNTIFS(\'{WB}\'!C2:C1000,"Open")',
     f'=COUNTIFS(\'{WA}\'!C2:C1000,"Open")',
     '=IF(B7>C7,"⬆",IF(B7<C7,"⬇","➡"))',None),
]
for row,label,fB,fA,fT,bg in kpi:
    ws.row_dimensions[row].height = 18
    sc(ws[f"A{row}"],label,bold=True,ha="left",bg=bg or "FFFFFFFF")
    sc(ws[f"B{row}"],fB,ha="center",bg=bg or "FFFFFFFF",fmt="0%" if row==6 else "General")
    sc(ws[f"C{row}"],fA,ha="center",bg=bg or "FFFFFFFF",fmt="0%" if row==6 else "General")
    sc(ws[f"D{row}"],fT,sz=14,ha="center",bg=bg or "FFFFFFFF")

for coord,val,ha in [("A9","Query Category","left"),("B9","Received","center"),
                      ("C9","Closed","center"),("D9","Closed Within TAT","center"),("E9","TAT %","center")]:
    sc(ws[coord],val,bold=True,bg=PURPLE,fc="FFFFFFFF",ha=ha)

CATS = sorted(excl_df["Query Category"].dropna().unique().tolist())
if excl_df["Query Category"].isna().any(): CATS.append("(Blank)")

for i,cat in enumerate(CATS):
    r = 10+i; ws.row_dimensions[r].height = 18
    sc(ws[f"A{r}"],cat,bold=True,ha="left")
    if cat=="(Blank)":
        ws[f"B{r}"].value=ArrayFormula(f"B{r}",f'=SUMPRODUCT(({EX}!H2:H1000="")*(({EX}!H2:H1000)<>FALSE)*1)')
        ws[f"B{r}"].font=Font(size=11,name="Calibri"); ws[f"B{r}"].alignment=Alignment(horizontal="center",vertical="center")
        sc(ws[f"C{r}"],f'=COUNTIFS({EX}!C2:C1000,"Closed",{EX}!H2:H1000,"")',ha="center")
        sc(ws[f"D{r}"],f'=COUNTIFS({EX}!C2:C1000,"Closed",{EX}!H2:H1000,"",{EX}!A2:A1000,"Not Violated")',ha="center")
    else:
        sc(ws[f"B{r}"],f'=COUNTIF({EX}!H2:H1000,"{cat}")',ha="center")
        sc(ws[f"C{r}"],f'=COUNTIFS({EX}!H2:H1000,"{cat}",{EX}!C2:C1000,"Closed")',ha="center")
        sc(ws[f"D{r}"],f'=COUNTIFS({EX}!H2:H1000,"{cat}",{EX}!C2:C1000,"Closed",{EX}!A2:A1000,"Not Violated")',ha="center")
    sc(ws[f"E{r}"],f"=IFERROR(D{r}/C{r},0)",bold=True,ha="center",fmt="0%")

buf = io.BytesIO(); wb.save(buf); buf.seek(0); xl_bytes = buf.read()
print("Excel built.")

# ── Send email ────────────────────────────────────────────────
today_str = datetime.now().strftime("%d %b %Y")
fname     = f"TAT_Report_{WK_B.replace(' ','_')}_{WK_A.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
subject   = f"Valuecart Seller Queries TAT Report – {WK_B} & {WK_A} ({today_str})"
body      = f"""Hi Team,

Please find the Valuecart Seller Queries TAT Report for the last two completed weeks.

{WK_B}: {wB_start.strftime('%d %b')} – {wB_end.strftime('%d %b %Y')}
{WK_A}: {wA_start.strftime('%d %b')} – {wA_end.strftime('%d %b %Y')}

Regards"""

msg = MIMEMultipart()
msg["from"]    = GMAIL_ADDRESS
msg["to"]      = EMAIL_TO
msg["subject"] = subject
if EMAIL_CC: msg["cc"] = EMAIL_CC
msg.attach(MIMEText(body,"plain"))

att = MIMEBase("application","vnd.openxmlformats-officedocument.spreadsheetml.sheet")
att.set_payload(xl_bytes); encoders.encode_base64(att)
att.add_header("Content-Disposition","attachment",filename=fname)
msg.attach(att)

recipients = [e.strip() for e in EMAIL_TO.split(",")]
if EMAIL_CC: recipients += [e.strip() for e in EMAIL_CC.split(",")]

with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
    s.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
    s.sendmail(GMAIL_ADDRESS, recipients, msg.as_bytes())
print(f"Email sent to {recipients}")
